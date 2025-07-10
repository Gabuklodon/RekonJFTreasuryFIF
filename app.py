import re
import pandas as pd
import os
import logging
from flask import Flask, request, send_file, render_template, jsonify, session
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from werkzeug.utils import secure_filename
import shutil
import uuid
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your-secret-key-here-change-in-production'  # Change this in production

# Configuration
UPLOAD_FOLDER = 'uploads'
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB
ALLOWED_EXTENSIONS = {'txt'}
MAX_FILES = 50

# Define column template with improved structure
COLUMNS = [
    "NO",
    "BANK JF/SOFCODE",
    "Dana Pembayaran Jumlah",
    "Pembayaran Angsuran Jumlah", "Pembayaran Angsuran Acc",
    "Pembayaran Denda Jumlah", "Pembayaran Denda Acc",
    "Pelunasan dipercepat Jumlah", "Pelunasan dipercepat Acc",
    "Denda Pelunasan dipercepat Jumlah", "Denda Pelunasan dipercepat Acc",
    "Penalti Pelunasan dipercepat Jumlah", "Penalti Pelunasan dipercepat Acc",
    "Pelunasan dipercepat case Asuransi Jumlah", "Pelunasan dipercepat case Asuransi Acc",
    "Penalti pelunasan dipercepat case Asuransi Jumlah", "Penalti pelunasan dipercepat case Asuransi Acc",
    "Pembayaran Recovery Jumlah", "Pembayaran Recovery Acc",
    "Penghapusan denda konsumen Jumlah", "Penghapusan denda konsumen Acc",
]

def extract_code_from_filename(filename: str) -> str:
    """Extract BANK JF/SOFCODE from the uploaded filename.

    * Accepts patterns such as 'JFCS2COVI-1_FIFJIN_YYMMDD.txt'.
    * Removes trailing COVID/COVI suffixes + digits, maps legacy *JFJ* ➜ *JASA*.
    """
    try:
        base = os.path.splitext(os.path.basename(filename))[0]          # JFCS2COVI-1_...
        first = re.split(r"[-_]", base, maxsplit=1)[0].upper()         # JFCS2COVI
        first = re.sub(r"(COVID|COVI)\d*$", "", first, flags=re.I)   # JFCS2
        return "JASA" if first == "JFJ" else first
    except Exception as exc:
        logger.error("[extract_code_from_filename] %s: %s", filename, exc)
        return ""

    """
    Extract SOFCODE from filename with specific rules.
    
    Expected filename format: SOFCODE-1_FIFIP_PKKF01072025.txt
    Special rules:
    - JFJ -> JASA (only special mapping)
    - JFJR -> JFJR (keep as-is)
    - JFPI5COVID -> JFPI5 (shortened)
    - Others -> keep as-is
    """
    try:
        # Remove file extension
        name_without_ext = filename.replace('.txt', '').replace('.TXT', '')
        
        # Extract the part before first "-" or "_" 
        if '-' in name_without_ext:
            code_part = name_without_ext.split('-')[0]
        elif '_' in name_without_ext:
            code_part = name_without_ext.split('_')[0]
        else:
            code_part = name_without_ext
        
        # Clean the code part
        code_part = code_part.strip().upper()
        
        # Only JFJ gets special mapping to JASA
        if code_part == 'JFJ':
            logger.info(f"Applied special mapping: {code_part} -> JASA")
            return 'JASA'
        
        # Special handling for JFPI5COVID - shorten to JFPI5
        if code_part == 'JFPI5COVID':
            result = 'JFPI5'
            logger.info(f"Shortened JFPI5COVID to: {result}")
            return result
        
        # For all other cases (including JFJR), return the extracted code as-is
        if code_part and len(code_part) >= 2:  # Ensure it's a valid code
            logger.info(f"Using filename code as-is: {code_part}")
            return code_part
        
        return ""
        
    except Exception as e:
        logger.error(f"Error extracting code from filename {filename}: {str(e)}")
        return ""

def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validate_file_size(file):
    """Validate file size."""
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    return size <= MAX_FILE_SIZE

def extract_text_from_file(file_bytes: bytes) -> str:
    """Decode TXT bytes and strip non-printable characters with better error handling."""
    try:
        # Try UTF-8 first
        text = file_bytes.decode('utf-8', errors='ignore')
    except UnicodeDecodeError:
        try:
            # Fallback to latin-1
            text = file_bytes.decode('latin-1', errors='ignore')
        except UnicodeDecodeError:
            # Last resort - CP1252 (Windows)
            text = file_bytes.decode('cp1252', errors='ignore')
    
    # Clean non-printable characters but keep newlines and tabs
    text = re.sub(r'[^\x20-\x7E\n\t]', '', text)
    return text

def to_int(num_str: str) -> int:
    """Convert a formatted number string (e.g. '1,057' or '1.234,56') to int.

    Strips every non‑digit character so that any thousands/decimal separators or
    hidden control characters (e.g. ESC) are ignored.
    """
    if not num_str:
        return 0
    cleaned = re.sub(r"[^\d]", "", num_str)
    return int(cleaned) if cleaned else 0

def sum_category(keyword_pattern: str, text: str) -> tuple[int, int]:
    """
    Cari '<keyword_pattern> ... Rp. <amount> ... untuk <count> Konsumen'
    • Kompatibel dengan karakter kontrol E F, spasi ganda, dsb.
    • Jika amount == 0 → paksa count = 0
    """
    pat = (
        rf"{keyword_pattern}"                       # kata kunci
        rf".*?Rp\.\s*[^\d]*"                       # Rp. ... lalu bebas (non‐digit)          
        rf"(?P<amt>[\d\.,]+)"                      #    ← amount (boleh ada , .)
        rf".*?untuk\s+[^\d]*"                      # sampai kata 'untuk'
        rf"(?P<cnt>[\d\.,]+)"                      #    ← count (boleh , .)
        rf"[^\d]*\s+Konsumen"                      # sebelum kata Konsumen
    )
    m = re.search(pat, text, flags=re.I | re.S)
    if not m:
        return 0, 0

    amt = to_int(m.group("amt"))
    cnt = to_int(m.group("cnt"))

    # --- aturan bisnis: jika amount = 0, acc harus 0 ---
    if amt == 0:
        cnt = 0

    return amt, cnt

def parse_jf_text(file_bytes: bytes, filename: str = "") -> dict:
    """Parse one JF *.txt* file into a structured dict ready for DataFrame."""

    # NOTE: helper `extract_text_from_file` & constant `COLUMNS` are assumed
    #       to already exist in the original codebase.

    try:
        text = extract_text_from_file(file_bytes)
        data = {"filename": filename}

        # --- init all fields with safe defaults --------------------------------
        for col in COLUMNS[1:]:
            data[col] = 0 if ("Jumlah" in col or "Acc" in col) else ""

        # --- BANK JF/SOFCODE ----------------------------------------------------
        code_from_file = extract_code_from_filename(filename)
        if code_from_file:
            data["BANK JF/SOFCODE"] = code_from_file
        else:
            for pat in [
                r"\(\s*([A-Z0-9]+)\s*\)",
                r"SOFCODE\s*[:\s]*([A-Z0-9]+)",
                r"Bank\s+[^\(]*\(\s*([A-Z0-9]+)\s*\)",
            ]:
                m = re.search(pat, text, flags=re.I)
                if m:
                    data["BANK JF/SOFCODE"] = m.group(1).upper()
                    break

        # --- Dana Pembayaran ----------------------------------------------------
        for pat in [
            r"sejumlah\s+Rp\.\s*[^\d]*(?P<amt>[\d\.,]+)",
            r"jumlah\s+Rp\.\s*[^\d]*(?P<amt>[\d\.,]+)",
            r"sebesar\s+Rp\.\s*[^\d]*(?P<amt>[\d\.,]+)",
        ]:
            m = re.search(pat, text, flags=re.I)
            if m:
                data["Dana Pembayaran Jumlah"] = to_int(m.group("amt"))
                break

        # --- Angsuran (poin a) --------------------------------------------------
        m = re.search(
            r"a\)\s*Pembayaran\s+angsuran\s+sebesar\s+Rp\.\s*[^\d]*(?P<amt>[\d\.,]+)"  # amount
            r".*?untuk\s+[^\d]*(?P<cnt>[\d\.,]+)[^\d]*\s+Konsumen",                       # count
            text,
            flags=re.I | re.S,
        )
        if m:
            data["Pembayaran Angsuran Jumlah"] = to_int(m.group("amt"))
            data["Pembayaran Angsuran Acc"]    = to_int(m.group("cnt"))

        # --- Remaining categories (b, c, d, ...) --------------------------------
        categories = [
            ("Pembayaran Denda",                       r"Pembayaran\s+denda"),
            ("Pelunasan dipercepat",                   r"Pembayaran\s+pelunasan\s+dipercepat"),
            ("Denda Pelunasan dipercepat",             r"Denda\s+pelunasan\s+dipercepat"),
            ("Penalti Pelunasan dipercepat",           r"Pembayaran\s+penalti\s+pelunasan\s+dipercepat"),
            ("Pelunasan dipercepat case Asuransi",     r"pelunasan\s+dipercepat\s+karena\s+pencairan\s+tagihan\s+asuransi"),
            ("Penalti pelunasan dipercepat case Asuransi", r"penalti\s+pelunasan\s+dipercepat\s+karena\s+pencairan\s+tagihan\s+asuransi"),
            ("Pembayaran Recovery",                    r"Pembayaran\s+recovery"),
            ("Penghapusan denda konsumen",             r"Penghapusan\s+denda"),
        ]
        for label, kw_pat in categories:
            amt, cnt = sum_category(kw_pat, text)
            data[f"{label} Jumlah"] = amt
            data[f"{label} Acc"]    = cnt

        return data

    except Exception as exc:
        logger.exception("Error processing file %s", filename)
        data = {"filename": filename, "error": str(exc)}
        for col in COLUMNS[1:]:
            data[col] = 0 if ("Jumlah" in col or "Acc" in col) else ""
        return data

    """Parse JF text with enhanced error handling and logging."""
    try:
        text = extract_text_from_file(file_bytes)
        data = {"filename": filename}
        
        logger.info(f"Processing file: {filename}")
        
        # Initialize all fields with default values first
        for col in COLUMNS[1:]:
            if "Jumlah" in col or "Acc" in col:
                data[col] = 0
            else:
                data[col] = ""
        
        code_from_file = extract_code_from_filename(filename)
        if code_from_file:
            data["BANK JF/SOFCODE"] = code_from_file
            logger.info(f"SOFCODE dari filename: {code_from_file} ({filename})")
        else:
            # Fallback ke isi teks jika pola nama file tidak terbaca
            patterns_bank = [
                r"\(\s*([A-Z0-9]+)\s*\)",                      # (JFCS2)
                r"SOFCODE\s*[:\s]*([A-Z0-9]+)",
                r"Bank\s+[^\(]*\(\s*([A-Z0-9]+)\s*\)",         # “Bank XXX (JFCS2)”
            ]
            for pat in patterns_bank:
                m = re.search(pat, text, flags=re.I)
                if m:
                    data["BANK JF/SOFCODE"] = m.group(1).upper()
                    logger.info(f"SOFCODE dari teks: {m.group(1)} ({filename})")
                    break
            else:
                logger.warning(f"SOFCODE tidak ditemukan dalam teks maupun nama file {filename}")
                data["BANK JF/SOFCODE"] = ""

        # Extract Dana Pembayaran with multiple patterns
        patterns_dana = [
            r"sejumlah\s+Rp\.\s*([^\d]*(?P<amt>[\d\.,]+))",
            r"jumlah\s+Rp\.\s*([^\d]*(?P<amt>[\d\.,]+))",
            r"sebesar\s+Rp\.\s*([^\d]*(?P<amt>[\d\.,]+))"
        ]
        
        for pattern in patterns_dana:
            m = re.search(pattern, text, flags=re.IGNORECASE)
            if m:
                try:
                    data["Dana Pembayaran Jumlah"] = to_int(m.group('amt'))
                    break
                except:
                    continue

        # Extract Pembayaran Angsuran (point a) with simpler approach
        try:
            pattern_angsuran = (
                r"a\)\s*Pembayaran\s+angsuran\s+sebesar\s+Rp\.\s*[^\d]*"
                r"(?P<amt>[\d\.,]+).*?untuk\s+[^\d]*(?P<cnt>\d+)[^\d]*\s+Konsumen"
            )
            m = re.search(pattern_angsuran, text, flags=re.IGNORECASE|re.DOTALL)
            if m:
                data["Pembayaran Angsuran Jumlah"] = to_int(m.group('amt'))
                data["Pembayaran Angsuran Acc"] = int(m.group('cnt'))
        except Exception as e:
            logger.warning(f"Error parsing angsuran in {filename}: {str(e)}")

        # Extract other categories using simplified sum_category function
        categories = [
            ("Pembayaran Denda", r"Pembayaran\s+denda"),
            ("Pelunasan dipercepat", r"Pembayaran\s+pelunasan\s+dipercepat"),
            ("Denda Pelunasan dipercepat", r"Denda\s+pelunasan\s+dipercepat"),
            ("Penalti Pelunasan dipercepat", r"Pembayaran\s+penalti\s+pelunasan\s+dipercepat"),
            ("Pelunasan dipercepat case Asuransi", r"pelunasan\s+dipercepat\s+karena\s+pencairan\s+tagihan\s+asuransi"),
            ("Penalti pelunasan dipercepat case Asuransi", r"penalti\s+pelunasan\s+dipercepat\s+karena\s+pencairan\s+tagihan\s+asuransi"),
            ("Pembayaran Recovery", r"Pembayaran\s+recovery"),
            ("Penghapusan denda konsumen", r"Penghapusan\s+denda"),
        ]
        
        for category_name, pattern in categories:
            try:
                amt, cnt = sum_category(pattern, text)
                data[f"{category_name} Jumlah"] = amt
                data[f"{category_name} Acc"] = cnt
            except Exception as e:
                logger.warning(f"Error parsing {category_name} in {filename}: {str(e)}")
                data[f"{category_name} Jumlah"] = 0
                data[f"{category_name} Acc"] = 0

        logger.info(f"Successfully processed {filename}")
        return data
        
    except Exception as e:
        logger.error(f"Error processing file {filename}: {str(e)}")
        # Return safe data structure
        data = {"filename": filename, "error": str(e)}
        for col in COLUMNS[1:]:
            if "Jumlah" in col or "Acc" in col:
                data[col] = 0
            else:
                data[col] = ""
        return data

def create_excel_with_styling(df, output_path):
    """Create Excel file with enhanced styling and error handling."""
    try:
        # Write DataFrame to Excel first
        logger.info("Creating Excel file...")
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, header=False, startrow=2)

        # Load workbook for styling
        wb = load_workbook(output_path)
        ws = wb.active
        
        try:
            # Define styles with error handling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )
        except Exception as e:
            logger.warning(f"Error creating styles: {str(e)}. Using default styles.")
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            border = None
        
        # Header groups
        groups = [
            ("NO", 1),
            ("BANK JF/SOFCODE", 1),
            ("Dana Pembayaran", 1),
            ("Pembayaran Angsuran", 2),
            ("Pembayaran Denda", 2),
            ("Pelunasan dipercepat", 2),
            ("Denda Pelunasan dipercepat", 2),
            ("Penalti Pelunasan dipercepat", 2),
            ("Pelunasan case Asuransi", 2),
            ("Penalti case Asuransi", 2),
            ("Pembayaran Recovery", 2),
            ("Penghapusan denda konsumen", 2),
        ]

        # Create headers with error handling
        col = 1
        for title, span in groups:
            try:
                if span == 1:
                    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
                    cell = ws.cell(row=1, column=col, value=title)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if header_font:
                        cell.font = header_font
                    if header_fill:
                        cell.fill = header_fill
                    if border:
                        cell.border = border
                    col += 1
                else:
                    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+span-1)
                    cell = ws.cell(row=1, column=col, value=title)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if header_font:
                        cell.font = header_font
                    if header_fill:
                        cell.fill = header_fill
                    if border:
                        cell.border = border
                    
                    # Sub headers
                    subcell1 = ws.cell(row=2, column=col, value="Jumlah")
                    subcell1.alignment = Alignment(horizontal="center", vertical="center")
                    if header_font:
                        subcell1.font = header_font
                    if header_fill:
                        subcell1.fill = header_fill
                    if border:
                        subcell1.border = border
                    
                    subcell2 = ws.cell(row=2, column=col+1, value="Acc")
                    subcell2.alignment = Alignment(horizontal="center", vertical="center")
                    if header_font:
                        subcell2.font = header_font
                    if header_fill:
                        subcell2.fill = header_fill
                    if border:
                        subcell2.border = border
                    
                    col += span
            except Exception as e:
                logger.warning(f"Error creating header for {title}: {str(e)}")
                col += span  # Continue with next column

        # Style data cells with error handling
        try:
            for row in range(3, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    try:
                        cell = ws.cell(row=row, column=col)
                        if border:
                            cell.border = border
                        if col > 3:  # Numeric columns
                            cell.alignment = Alignment(horizontal="right")
                        else:
                            cell.alignment = Alignment(horizontal="center")
                    except Exception as e:
                        logger.warning(f"Error styling cell ({row}, {col}): {str(e)}")
                        continue
        except Exception as e:
            logger.warning(f"Error styling data cells: {str(e)}")

        # Auto-adjust column widths with error handling
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 30)  # Min 10, Max 30
                ws.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            logger.warning(f"Error adjusting column widths: {str(e)}")

        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel file created successfully: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Critical error creating Excel file: {str(e)}")
        # Try to create a basic Excel file without styling as fallback
        try:
            logger.info("Attempting to create basic Excel file as fallback...")
            df.to_excel(output_path, index=False)
            logger.info("Basic Excel file created successfully")
            return True
        except Exception as fallback_error:
            logger.error(f"Fallback Excel creation also failed: {str(fallback_error)}")
            return False

def process_files(upload_folder, output_path):
    """Process uploaded files with enhanced error handling."""
    try:
        rows = []
        no = 1
        processed_files = 0
        error_files = []
        
        for filename in os.listdir(upload_folder):
            if filename.endswith('.txt'):
                file_path = os.path.join(upload_folder, filename)
                try:
                    with open(file_path, 'rb') as f:
                        raw = f.read()
                    
                    data = parse_jf_text(raw, filename)
                    
                    if "error" in data:
                        error_files.append(filename)
                        logger.warning(f"Error in file {filename}: {data['error']}")
                    
                    row = [no] + [data.get(col, "") for col in COLUMNS[1:]]
                    rows.append(row)
                    processed_files += 1
                    no += 1
                    
                except Exception as e:
                    error_files.append(filename)
                    logger.error(f"Failed to process file {filename}: {str(e)}")

        if not rows:
            logger.error("No valid data found in any files")
            return False, "No valid data found in uploaded files"

        # Create DataFrame
        df = pd.DataFrame(rows, columns=COLUMNS)
        
        # Create Excel with styling
        success = create_excel_with_styling(df, output_path)
        
        if success:
            message = f"Successfully processed {processed_files} files"
            if error_files:
                message += f" ({len(error_files)} files had errors)"
            return True, message
        else:
            return False, "Failed to create Excel file"
            
    except Exception as e:
        logger.error(f"Error in process_files: {str(e)}")
        return False, f"Processing error: {str(e)}"

@app.route('/')
def index():
    """Serve the main page and clear session."""
    session.clear()
    # Return the HTML content directly instead of using template
    return '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Rekon JF Processor - FIFGroup</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
        body { font-family: 'Inter', sans-serif; }
        
        .gradient-bg {
            background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 35%, #60a5fa 70%, #93c5fd 100%);
        }
        
        .fifgroup-gradient {
            background: linear-gradient(135deg, #1e40af 0%, #2563eb 50%, #3b82f6 100%);
        }
        
        .fifgroup-primary { background-color: #1e40af; }
        .fifgroup-secondary { background-color: #2563eb; }
        .fifgroup-accent { background-color: #3b82f6; }
        
        .glass-effect {
            background: rgba(255, 255, 255, 0.98);
            backdrop-filter: blur(15px);
            border: 1px solid rgba(59, 130, 246, 0.2);
            box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.15);
        }
        
        .file-drop-zone {
            border: 2px dashed #3b82f6;
            transition: all 0.3s ease;
            background: linear-gradient(145deg, #f8fafc 0%, #f1f5f9 100%);
        }
        
        .file-drop-zone.dragover {
            border-color: #1e40af;
            background: linear-gradient(145deg, #eff6ff 0%, #dbeafe 100%);
            transform: scale(1.02);
        }
        
        .progress-bar {
            transition: width 0.4s ease;
            background: linear-gradient(90deg, #1e40af 0%, #2563eb 50%, #3b82f6 100%);
        }
        
        .fifgroup-btn-primary {
            background: linear-gradient(135deg, #1e40af 0%, #2563eb 100%);
            transition: all 0.3s ease;
        }
        
        .fifgroup-btn-primary:hover {
            background: linear-gradient(135deg, #1e3a8a 0%, #1d4ed8 100%);
            transform: translateY(-2px);
            box-shadow: 0 10px 25px -5px rgba(30, 64, 175, 0.4);
        }
        
        .fifgroup-btn-process {
            background: linear-gradient(135deg, #2563eb 0%, #3b82f6 100%);
        }
        
        .fifgroup-btn-process:hover {
            background: linear-gradient(135deg, #1d4ed8 0%, #2563eb 100%);
            transform: translateY(-2px);
            box-shadow: 0 10px 25px -5px rgba(37, 99, 235, 0.4);
        }
        
        .fifgroup-btn-download {
            background: linear-gradient(135deg, #059669 0%, #10b981 100%);
        }
        
        .fifgroup-btn-download:hover {
            background: linear-gradient(135deg, #047857 0%, #059669 100%);
            transform: translateY(-2px);
            box-shadow: 0 10px 25px -5px rgba(5, 150, 105, 0.4);
        }
        
        .corporate-header {
            background: linear-gradient(135deg, #1e40af 0%, #3b82f6 50%, #60a5fa 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }
        
        .pulse-dot {
            animation: pulse 2s cubic-bezier(0.4, 0, 0.6, 1) infinite;
        }
        
        @keyframes pulse {
            0%, 100% { opacity: 1; }
            50% { opacity: 0.5; }
        }
        
        .file-item {
            animation: slideIn 0.3s ease-out;
            background: linear-gradient(145deg, #f8fafc 0%, #f1f5f9 100%);
            border: 1px solid rgba(59, 130, 246, 0.1);
        }
        
        @keyframes slideIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        .success-checkmark {
            animation: checkmark 0.6s ease-in-out;
        }
        
        @keyframes checkmark {
            0% { transform: scale(0); }
            50% { transform: scale(1.2); }
            100% { transform: scale(1); }
        }
        
        .step-active {
            background: linear-gradient(135deg, #1e40af 0%, #2563eb 100%);
            box-shadow: 0 0 20px rgba(30, 64, 175, 0.5);
        }
        
        .step-completed {
            background: linear-gradient(135deg, #059669 0%, #10b981 100%);
            box-shadow: 0 0 20px rgba(5, 150, 105, 0.5);
        }
        
        .step-inactive {
            background: #e5e7eb;
            color: #9ca3af;
        }
        
        .fifgroup-icon {
            background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
        }
    </style>
</head>
<body class="gradient-bg min-h-screen flex items-center justify-center p-4">
    <div class="glass-effect rounded-2xl shadow-2xl w-full max-w-2xl p-8">
        <!-- Header with FIFGroup Branding -->
        <div class="text-center mb-8">
            <div class="inline-flex items-center justify-center w-20 h-20 fifgroup-icon rounded-full mb-4 shadow-lg">
                <i class="fas fa-file-excel text-white text-3xl"></i>
            </div>
            <h1 class="text-4xl font-bold corporate-header mb-2">Rekon JF Processor</h1>
            <div class="text-sm text-blue-600 font-semibold mb-2">FIFGroup Finance & Treasury Division</div>
            <p class="text-gray-600">Upload your TXT files and convert them to structured Excel format</p>
            <div class="w-16 h-1 fifgroup-gradient mx-auto mt-3 rounded-full"></div>
        </div>

        <!-- Progress Steps with FIFGroup Styling -->
        <div class="flex justify-between mb-8">
            <div class="flex items-center">
                <div id="step1" class="w-10 h-10 step-active rounded-full flex items-center justify-center text-white text-sm font-bold shadow-lg">1</div>
                <span class="ml-3 text-sm font-semibold text-gray-700">Upload</span>
            </div>
            <div class="flex-1 h-2 mx-4 bg-gray-200 rounded-full">
                <div id="progress1" class="h-full rounded-full progress-bar w-0"></div>
            </div>
            <div class="flex items-center">
                <div id="step2" class="w-10 h-10 step-inactive rounded-full flex items-center justify-center text-sm font-bold">2</div>
                <span class="ml-3 text-sm font-medium text-gray-500">Process</span>
            </div>
            <div class="flex-1 h-2 mx-4 bg-gray-200 rounded-full">
                <div id="progress2" class="h-full rounded-full progress-bar w-0"></div>
            </div>
            <div class="flex items-center">
                <div id="step3" class="w-10 h-10 step-inactive rounded-full flex items-center justify-center text-sm font-bold">3</div>
                <span class="ml-3 text-sm font-medium text-gray-500">Download</span>
            </div>
        </div>

        <!-- File Upload Area with Enhanced Styling -->
        <div id="uploadArea" class="file-drop-zone rounded-xl p-8 text-center mb-6 shadow-inner">
            <input type="file" id="fileInput" multiple accept=".txt" class="hidden">
            <div class="mb-4">
                <i class="fas fa-cloud-upload-alt text-5xl text-blue-500 mb-4"></i>
                <p class="text-xl font-semibold text-gray-700 mb-2">Drop your TXT files here</p>
                <p class="text-sm text-gray-500 mb-6">or click to browse • Maximum 50 files, 16MB each</p>
                <button onclick="document.getElementById('fileInput').click()" 
                        class="fifgroup-btn-primary text-white px-8 py-3 rounded-lg font-semibold shadow-lg disabled:opacity-50 disabled:cursor-not-allowed">
                    <i class="fas fa-folder-open mr-2"></i>
                    Choose Files
                </button>
            </div>
        </div>

        <!-- Selected Files with Enhanced Design -->
        <div id="filesList" class="hidden mb-6">
            <h3 class="text-lg font-bold text-gray-800 mb-4 flex items-center">
                <i class="fas fa-list-ul text-blue-600 mr-2"></i>
                Selected Files
            </h3>
            <div id="filesContainer" class="space-y-3 max-h-40 overflow-y-auto"></div>
        </div>

        <!-- Action Buttons with FIFGroup Styling -->
        <div class="space-y-4">
            <button id="uploadBtn" onclick="uploadFiles()" class="w-full fifgroup-btn-primary text-white py-4 px-6 rounded-xl font-semibold shadow-lg disabled:opacity-50 disabled:cursor-not-allowed disabled:transform-none">
                <i class="fas fa-upload mr-3"></i>
                Upload Files
            </button>
            
            <button id="processBtn" onclick="processFiles()" disabled 
                    class="w-full fifgroup-btn-process text-white py-4 px-6 rounded-xl font-semibold shadow-lg disabled:opacity-50 disabled:cursor-not-allowed disabled:transform-none">
                <i class="fas fa-cogs mr-3"></i>
                Process Files
            </button>
            
            <button id="downloadBtn" onclick="downloadFile()" disabled 
                    class="w-full fifgroup-btn-download text-white py-4 px-6 rounded-xl font-semibold shadow-lg disabled:opacity-50 disabled:cursor-not-allowed disabled:transform-none">
                <i class="fas fa-download mr-3"></i>
                Download Excel
            </button>
        </div>

        <!-- Status Area with Enhanced Design -->
        <div id="statusArea" class="mt-8">
            <div id="status" class="text-center text-gray-600 font-medium"></div>
            <div id="loadingSpinner" class="hidden flex justify-center items-center mt-4">
                <div class="animate-spin rounded-full h-10 w-10 border-b-2 border-blue-600"></div>
                <span class="ml-3 text-gray-700 font-medium">Processing...</span>
            </div>
        </div>

        <!-- Success Message with FIFGroup Styling -->
        <div id="successMessage" class="hidden mt-6 p-4 bg-gradient-to-r from-green-50 to-emerald-50 border border-green-200 text-green-800 rounded-xl shadow-sm">
            <div class="flex items-center">
                <i class="fas fa-check-circle success-checkmark mr-3 text-green-600"></i>
                <span class="font-semibold">Files processed successfully! Ready to download.</span>
            </div>
        </div>
        
        <!-- Error Details with Enhanced Styling -->
        <div id="errorDetails" class="hidden mt-6 p-4 bg-gradient-to-r from-red-50 to-pink-50 border border-red-200 text-red-800 rounded-xl shadow-sm">
            <div class="flex items-center mb-2">
                <i class="fas fa-exclamation-triangle mr-3 text-red-600"></i>
                <span class="font-semibold">Error Details:</span>
            </div>
            <div id="errorMessage" class="text-sm pl-6"></div>
        </div>
        
        <!-- Footer with Company Info -->
        <div class="mt-8 pt-6 border-t border-gray-200 text-center">
            <p class="text-xs text-gray-500">
                <i class="fas fa-shield-alt mr-1"></i>
                Powered by FIFGroup Member of ASTRA
            </p>
        </div>
    </div>

    <script>
        let selectedFiles = [];
        
        // File input change handler
        document.getElementById('fileInput').addEventListener('change', function(e) {
            handleFiles(e.target.files);
        });

        // Drag and drop handlers
        const uploadArea = document.getElementById('uploadArea');
        
        uploadArea.addEventListener('dragover', function(e) {
            e.preventDefault();
            uploadArea.classList.add('dragover');
        });
        
        uploadArea.addEventListener('dragleave', function(e) {
            e.preventDefault();
            uploadArea.classList.remove('dragover');
        });
        
        uploadArea.addEventListener('drop', function(e) {
            e.preventDefault();
            uploadArea.classList.remove('dragover');
            handleFiles(e.dataTransfer.files);
        });

        function handleFiles(files) {
            selectedFiles = Array.from(files).filter(file => file.name.endsWith('.txt'));
            displaySelectedFiles();
            
            const uploadBtn = document.getElementById('uploadBtn');
            uploadBtn.disabled = selectedFiles.length === 0;
            
            if (selectedFiles.length === 0) {
                showStatus('Please select at least one TXT file.', 'error');
            } else {
                showStatus(`${selectedFiles.length} file(s) selected`, 'info');
            }
        }

        function displaySelectedFiles() {
            const filesList = document.getElementById('filesList');
            const filesContainer = document.getElementById('filesContainer');
            
            if (selectedFiles.length === 0) {
                filesList.classList.add('hidden');
                return;
            }
            
            filesList.classList.remove('hidden');
            filesContainer.innerHTML = '';
            
            selectedFiles.forEach((file, index) => {
                const fileItem = document.createElement('div');
                fileItem.className = 'file-item flex items-center justify-between p-4 rounded-xl shadow-sm border';
                fileItem.innerHTML = `
                    <div class="flex items-center">
                        <i class="fas fa-file-alt text-blue-600 mr-4 text-lg"></i>
                        <div>
                            <p class="font-semibold text-gray-800">${file.name}</p>
                            <p class="text-sm text-gray-500">${formatFileSize(file.size)}</p>
                        </div>
                    </div>
                    <button onclick="removeFile(${index})" class="text-red-500 hover:text-red-700 p-2 rounded-lg hover:bg-red-50 transition-colors">
                        <i class="fas fa-times"></i>
                    </button>
                `;
                filesContainer.appendChild(fileItem);
            });
        }

        function removeFile(index) {
            selectedFiles.splice(index, 1);
            displaySelectedFiles();
            
            const uploadBtn = document.getElementById('uploadBtn');
            uploadBtn.disabled = selectedFiles.length === 0;
            
            if (selectedFiles.length === 0) {
                showStatus('No files selected', 'info');
            } else {
                showStatus(`${selectedFiles.length} file(s) selected`, 'info');
            }
        }

        function formatFileSize(bytes) {
            if (bytes === 0) return '0 Bytes';
            const k = 1024;
            const sizes = ['Bytes', 'KB', 'MB', 'GB'];
            const i = Math.floor(Math.log(bytes) / Math.log(k));
            return parseFloat((bytes / Math.pow(k, i)).toFixed(2)) + ' ' + sizes[i];
        }

        function updateStep(step, completed = false) {
            const stepElement = document.getElementById(`step${step}`);
            const progressElement = document.getElementById(`progress${step}`);
            
            if (completed) {
                stepElement.className = 'w-10 h-10 step-completed rounded-full flex items-center justify-center text-white text-sm font-bold shadow-lg';
                stepElement.innerHTML = '<i class="fas fa-check"></i>';
                if (progressElement) {
                    progressElement.style.width = '100%';
                }
            } else {
                stepElement.className = 'w-10 h-10 step-active rounded-full flex items-center justify-center text-white text-sm font-bold shadow-lg';
                stepElement.textContent = step;
            }
            
            // Update next step to active if current is completed
            if (completed && step < 3) {
                const nextStep = document.getElementById(`step${step + 1}`);
                nextStep.className = 'w-10 h-10 step-active rounded-full flex items-center justify-center text-white text-sm font-bold shadow-lg';
                nextStep.textContent = step + 1;
            }
        }

        function showStatus(message, type = 'info') {
            const status = document.getElementById('status');
            const errorDetails = document.getElementById('errorDetails');
            
            const colors = {
                'info': 'text-blue-700 font-medium',
                'success': 'text-green-700 font-semibold',
                'error': 'text-red-700 font-semibold',
                'warning': 'text-amber-700 font-semibold'
            };
            
            status.className = `text-center ${colors[type] || colors.info}`;
            status.textContent = message;
            
            // Hide error details for non-error messages
            if (type !== 'error') {
                errorDetails.classList.add('hidden');
            }
        }

        function showError(message, details = null) {
            showStatus(message, 'error');
            
            if (details) {
                const errorDetails = document.getElementById('errorDetails');
                const errorMessage = document.getElementById('errorMessage');
                errorMessage.textContent = details;
                errorDetails.classList.remove('hidden');
            }
        }

        function showLoading(show = true) {
            const spinner = document.getElementById('loadingSpinner');
            if (show) {
                spinner.classList.remove('hidden');
            } else {
                spinner.classList.add('hidden');
            }
        }

        async function uploadFiles() {
            if (selectedFiles.length === 0) {
                showStatus('Please select at least one TXT file.', 'error');
                return;
            }

            const uploadBtn = document.getElementById('uploadBtn');
            const processBtn = document.getElementById('processBtn');
            
            uploadBtn.disabled = true;
            showLoading(true);
            showStatus('Uploading files...', 'info');

            const formData = new FormData();
            selectedFiles.forEach(file => {
                formData.append('files', file);
            });

            try {
                const response = await fetch('/upload', {
                    method: 'POST',
                    body: formData
                });
                const result = await response.json();

                showLoading(false);

                if (response.ok) {
                    updateStep(1, true);
                    updateStep(2, false);
                    showStatus(result.message, 'success');
                    processBtn.disabled = false;
                } else {
                    showError(`Upload failed: ${result.error}`);
                    uploadBtn.disabled = false;
                }
            } catch (error) {
                showLoading(false);
                showError('Network error during upload', error.message);
                uploadBtn.disabled = false;
            }
        }

        async function processFiles() {
            const processBtn = document.getElementById('processBtn');
            const downloadBtn = document.getElementById('downloadBtn');
            
            processBtn.disabled = true;
            showLoading(true);
            showStatus('Processing files...', 'info');

            try {
                const response = await fetch('/process', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    }
                });
                
                const result = await response.json();
                showLoading(false);

                if (response.ok) {
                    updateStep(2, true);
                    updateStep(3, false);
                    showStatus(result.message, 'success');
                    downloadBtn.disabled = false;
                    document.getElementById('successMessage').classList.remove('hidden');
                } else {
                    showError(`Processing failed: ${result.error}`);
                    processBtn.disabled = false;
                }
            } catch (error) {
                showLoading(false);
                showError('Network error during processing', error.message);
                processBtn.disabled = false;
            }
        }

        async function downloadFile() {
            const downloadBtn = document.getElementById('downloadBtn');
            const uploadBtn = document.getElementById('uploadBtn');
            
            downloadBtn.disabled = true;
            showLoading(true);
            showStatus('Preparing download...', 'info');

            try {
                const response = await fetch('/download');
                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = 'rekon_jf.xlsx';
                    document.body.appendChild(a);
                    a.click();
                    a.remove();
                    window.URL.revokeObjectURL(url);

                    showLoading(false);
                    updateStep(3, true);
                    showStatus('File downloaded successfully! Ready for new files.', 'success');
                    
                    // Reset for new upload
                    setTimeout(() => {
                        resetForm();
                    }, 2000);
                } else {
                    const result = await response.json();
                    showLoading(false);
                    showError(`Download failed: ${result.error}`);
                    downloadBtn.disabled = false;
                }
            } catch (error) {
                showLoading(false);
                showError('Network error during download', error.message);
                downloadBtn.disabled = false;
            }
        }

        function resetForm() {
            // Reset file selection
            selectedFiles = [];
            document.getElementById('fileInput').value = '';
            document.getElementById('filesList').classList.add('hidden');
            
            // Reset buttons
            document.getElementById('uploadBtn').disabled = false;
            document.getElementById('processBtn').disabled = true;
            document.getElementById('downloadBtn').disabled = true;
            
            // Reset steps with FIFGroup styling
            document.getElementById('step1').className = 'w-10 h-10 step-active rounded-full flex items-center justify-center text-white text-sm font-bold shadow-lg';
            document.getElementById('step1').textContent = '1';
            document.getElementById('step2').className = 'w-10 h-10 step-inactive rounded-full flex items-center justify-center text-sm font-bold';
            document.getElementById('step2').textContent = '2';
            document.getElementById('step3').className = 'w-10 h-10 step-inactive rounded-full flex items-center justify-center text-sm font-bold';
            document.getElementById('step3').textContent = '3';
            
            // Reset progress bars
            document.getElementById('progress1').style.width = '0%';
            document.getElementById('progress2').style.width = '0%';
            
            // Hide messages
            document.getElementById('successMessage').classList.add('hidden');
            document.getElementById('errorDetails').classList.add('hidden');
            
            // Reset status
            showStatus('Select TXT files to get started', 'info');
        }

        // Initialize
        showStatus('Select TXT files to get started', 'info');
    </script>
</body>
</html>'''

@app.route('/upload', methods=['POST'])
def upload_files():
    """Handle file upload with enhanced validation."""
    try:
        if 'files' not in request.files:
            return jsonify({'error': 'No files uploaded'}), 400

        files = request.files.getlist('files')
        if not files or all(f.filename == '' for f in files):
            return jsonify({'error': 'No files selected'}), 400

        # Validate number of files
        if len(files) > MAX_FILES:
            return jsonify({'error': f'Too many files. Maximum {MAX_FILES} files allowed'}), 400

        # Create unique upload folder
        upload_folder = os.path.join(UPLOAD_FOLDER, str(uuid.uuid4()))
        os.makedirs(upload_folder, exist_ok=True)

        uploaded_files = []
        invalid_files = []

        for file in files:
            if file and file.filename:
                # Validate file type
                if not allowed_file(file.filename):
                    invalid_files.append(f"{file.filename} (invalid type)")
                    continue
                
                # Validate file size
                if not validate_file_size(file):
                    invalid_files.append(f"{file.filename} (too large)")
                    continue

                filename = secure_filename(file.filename)
                if filename:  # Ensure filename is not empty after sanitization
                    file_path = os.path.join(upload_folder, filename)
                    file.save(file_path)
                    uploaded_files.append(filename)

        if not uploaded_files:
            shutil.rmtree(upload_folder, ignore_errors=True)
            error_msg = "No valid files uploaded"
            if invalid_files:
                error_msg += f". Issues: {', '.join(invalid_files)}"
            return jsonify({'error': error_msg}), 400

        session['upload_folder'] = upload_folder
        
        message = f"Successfully uploaded {len(uploaded_files)} file(s)"
        if invalid_files:
            message += f". Skipped {len(invalid_files)} invalid file(s)"
        
        logger.info(f"Upload successful: {message}")
        return jsonify({'message': message})

    except Exception as e:
        logger.error(f"Upload error: {str(e)}")
        return jsonify({'error': 'Upload failed due to server error'}), 500

@app.route('/process', methods=['POST'])
def process():
    """Process uploaded files with enhanced error handling."""
    try:
        upload_folder = session.get('upload_folder')
        if not upload_folder or not os.path.exists(upload_folder):
            return jsonify({'error': 'No uploaded files found. Please upload files first.'}), 400

        output_path = os.path.join(upload_folder, 'rekon_jf.xlsx')
        success, message = process_files(upload_folder, output_path)

        if not success:
            shutil.rmtree(upload_folder, ignore_errors=True)
            session.pop('upload_folder', None)
            return jsonify({'error': message}), 400

        session['output_path'] = output_path
        logger.info(f"Processing successful: {message}")
        return jsonify({'message': message})

    except Exception as e:
        logger.error(f"Processing error: {str(e)}")
        return jsonify({'error': 'Processing failed due to server error'}), 500

@app.route('/download')
def download():
    """Handle file download with cleanup."""
    try:
        output_path = session.get('output_path')
        if not output_path or not os.path.exists(output_path):
            return jsonify({'error': 'No processed file available. Please process files first.'}), 400

        upload_folder = os.path.dirname(output_path)
        
        # Generate timestamp for filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_name = f'rekon_jf_{timestamp}.xlsx'
        
        response = send_file(
            output_path, 
            as_attachment=True, 
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Schedule cleanup after response
        @response.call_on_close
        def cleanup():
            try:
                shutil.rmtree(upload_folder, ignore_errors=True)
                session.pop('upload_folder', None)
                session.pop('output_path', None)
                logger.info("Cleanup completed")
            except Exception as e:
                logger.error(f"Cleanup error: {str(e)}")

        logger.info(f"Download initiated: {download_name}")
        return response

    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        return jsonify({'error': 'Download failed due to server error'}), 500

@app.errorhandler(413)
def request_entity_too_large(error):
    """Handle file too large error."""
    return jsonify({'error': 'File too large. Maximum size is 16MB per file.'}), 413

@app.errorhandler(500)
def internal_server_error(error):
    """Handle internal server errors."""
    logger.error(f"Internal server error: {str(error)}")
    return jsonify({'error': 'Internal server error occurred'}), 500

if __name__ == '__main__':
    # Ensure upload directory exists
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    
    # Configure Flask for file uploads
    app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE * MAX_FILES
    
    logger.info("Starting Rekon JF Processor server...")
    app.run(debug=True, host='localhost', port=5000)